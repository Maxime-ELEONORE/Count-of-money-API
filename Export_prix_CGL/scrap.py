#! /usr/bin/env python3
from site_generation import get_all_site
from openpyxl.styles import PatternFill
from produit_class import Produit
import openpyxl
from datetime import datetime
#pour demain, créer une nouvelle variables dans la classe pour le title, pour pouvoir verifier si jamais la référence n'est pas la bonne
#pour demain, sur la référence, faut la nettoyer des espaces avant
#Prendre en compte les "/" dans les référence en fonction de chaque site comment il les transforment pour pouvoir s'en servir (le plus gros casse couilles c'est belong)
produits = []
workbook = openpyxl.load_workbook("Export prix CGL.xlsx")
sheet = workbook.active
# Parcourir chaque ligne à partir de la deuxième
for row in sheet.iter_rows(min_row=2, values_only=True):
    produits.append(Produit(row[2], row[0], row[3]))
# Fermer le classeur
workbook.close()
#produits.append(Produit("Hotte Cube ARTHUR MARTIN AFC4453MX", "AFC4453MX"))
#produits.append(Produit("Climatiseur mobile BEKO BPN112C", "BPN112C"))
#produits.append(Produit("table induction hotte", "cce84779cb"))
#produits.append(Produit("LAVE VAISSELLE FULL 42DB  ", "DSD444B/1"))
cnt = 0
sites = get_all_site()
for produit in produits:
    for site in sites:
        site_origin,price, url = site.get_price(produit.get_reference())
        produit.add_site(site_origin, price, url)
    cnt = cnt + 1
    if cnt== 50000:
        break
        """
        if price:
            print(f"Le prix du produit {produit.get_reference()} est: {price}€ sur {site_origin}")
        else :
            print(f"Le produit {produit.get_reference()} est introuvable sur: {site_origin}")
            """

# Créer un nouveau fichier Excel
workbook = openpyxl.Workbook()
sheet = workbook.active

# Obtenir la date et l'heure actuelles, formatées avec des "_"
current_time = datetime.now().strftime("%Y_%m_%d_%H_%M_%S")
file_name = f"resultat_{current_time}.xlsx"

# Écrire les valeurs dans la feuille à partir de la 2ème ligne
sheet.cell(row=1, column=1, value="Désignation")
sheet.cell(row=1, column=2, value="Référence")
sheet.cell(row=1, column=3, value="Notre prix")
sheet.cell(row=1, column=4, value="Url produit")

column_index_for_sites = 5
for site in sites:
    sheet.cell(row=1, column=column_index_for_sites, value=site.get_name())
    column_index_for_sites += 1

y = 2

blackFill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
redFill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
greenFill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

for produit in produits:
    sheet.cell(row=y, column=1, value=produit.get_name())
    sheet.cell(row=y, column=2, value=produit.get_reference())
    our_price_cell = sheet.cell(row=y, column=3, value=produit.get_ourprice())

    list_product_price = produit.get_price()
    list_product_url = produit.get_url_product()

    # Assurer que chaque produit a le même nombre de prix et d'URLs
    if len(list_product_price) != len(list_product_url):
        raise ValueError("Le nombre de prix ne correspond pas au nombre d'URLs pour un produit donné.")
    
    i = 5  # La colonne où les URLs commencent
    for url in list_product_url:
        sheet.cell(row=y, column=i, value=url)
        i += 1

    # Pas besoin d'incrémenter 'i' ici car nous ajoutons les prix immédiatement après les URLs
    for price in list_product_price:
        price_cell = sheet.cell(row=y, column=i, value=price)
        if price == -1:
            price_cell.fill = blackFill
        elif produit.get_ourprice() > price:
            price_cell.fill = redFill
        else:
            price_cell.fill = greenFill
        i += 1
    y += 1

# Sauvegarder le fichier Excel
workbook.save(file_name)
print(f"File saved as {file_name}")
